from typing import Any, Optional
from datetime import datetime, timezone, date
import io
import json
import logging

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import func, distinct
from sqlmodel import Session, select

from app.api.deps import get_current_user, get_current_distributor
from app.database import get_session
from app.models.distributor import Distributor
from app.models.objectif import Objectif
from app.models.user import User, UserRole
from app.models.vente import Vente

logger = logging.getLogger(__name__)

router = APIRouter()


@router.get("/next-missing")
def next_missing_month(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    current_distributor = Depends(get_current_distributor),
) -> Any:
    # Single query: get the latest (annee, mois) in one pass
    q = select(Objectif.annee, Objectif.mois).where(Objectif.mois.isnot(None), Objectif.annee.isnot(None))
    if current_user.role != UserRole.PLATFORM_ADMIN and current_distributor:
        q = q.where(Objectif.distributor_id == current_distributor.id)
    q = q.order_by(Objectif.annee.desc(), Objectif.mois.desc()).limit(1)
    row = session.exec(q).first()
    if not row:
        today = date.today()
        return {"mois": today.month, "annee": today.year}
    annee, mois = row
    if mois == 12:
        return {"mois": 1, "annee": annee + 1}
    return {"mois": mois + 1, "annee": annee}


@router.get("/routes-count")
def routes_count(
    mois: int = Query(..., ge=1, le=12),
    annee: int = Query(..., ge=2020),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    current_distributor = Depends(get_current_distributor),
) -> Any:
    annee_mois = f"{annee}-{mois:02d}"

    def _count_by_canal(am: str) -> dict:
        q = select(Vente.canal, func.count(distinct(Vente.code_fdv))).where(
            Vente.annee_mois == am,
            Vente.canal.in_(["VD", "VH"]),
            Vente.code_fdv.isnot(None),
        )
        if current_user.role != UserRole.PLATFORM_ADMIN and current_distributor:
            q = q.where((Vente.distributor_id == current_distributor.id) | (Vente.distributor_id == None))
        rows = session.exec(q.group_by(Vente.canal)).all()
        return {canal: int(cnt) for canal, cnt in rows}

    counts = _count_by_canal(annee_mois)
    vd = counts.get("VD", 0)
    vh = counts.get("VH", 0)

    fallback_mois = None
    if vd == 0 and vh == 0:
        q_latest = select(Vente.annee_mois).where(Vente.canal.isnot(None), Vente.code_fdv.isnot(None))
        if current_user.role != UserRole.PLATFORM_ADMIN and current_distributor:
            q_latest = q_latest.where((Vente.distributor_id == current_distributor.id) | (Vente.distributor_id == None))
        latest = session.exec(q_latest.order_by(Vente.annee_mois.desc()).limit(1)).first()
        if latest and latest != annee_mois:
            fallback_mois = latest
            fb = _count_by_canal(latest)
            vd = fb.get("VD", 0)
            vh = fb.get("VH", 0)

    return {"vd": vd, "vh": vh, "fallback_mois": fallback_mois}


@router.post("/batch")
def batch_upsert(
    mois: int = Query(..., ge=1, le=12),
    annee: int = Query(..., ge=2020),
    body: list = Body(...),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    current_distributor = Depends(get_current_distributor),
) -> Any:
    now = datetime.now(timezone.utc)

    def _f(item: dict, key: str) -> Optional[float]:
        v = item.get(key)
        return float(v) if v is not None else None

    # Resolve code_dd aliases → canonical code_produit
    # TODO: Produit model needs to be created for this to work
    dd_map = {}

    # Load all existing objectifs for this period in one query
    q = select(Objectif).where(Objectif.mois == mois, Objectif.annee == annee)
    if current_distributor:
        q = q.where(Objectif.distributor_id == current_distributor.id)
    existing = session.exec(q).all()
    obj_map = {o.code_produit: o for o in existing}

    for item in body:
        code = item.get("code_produit")
        if not code:
            continue
        code = dd_map.get(code, code)
        tonne_vd = _f(item, "objectif_tonne_vd")
        packs_vd = _f(item, "objectif_packs_vd")
        packs_vd_t = _f(item, "objectif_packs_vd_tournee")
        tonne_vh = _f(item, "objectif_tonne_vh")
        packs_vh = _f(item, "objectif_packs_vh")
        packs_vh_t = _f(item, "objectif_packs_vh_tournee")

        obj = obj_map.get(code)
        if obj:
            obj.objectif_tonne_vd = tonne_vd
            obj.objectif_packs_vd = packs_vd
            obj.objectif_packs_vd_tournee = packs_vd_t
            obj.objectif_tonne_vh = tonne_vh
            obj.objectif_packs_vh = packs_vh
            obj.objectif_packs_vh_tournee = packs_vh_t
            obj.updated_by_id = current_user.id
            obj.updated_at = now
        elif any(v is not None for v in (tonne_vd, packs_vd, packs_vd_t, tonne_vh, packs_vh, packs_vh_t)):
            session.add(Objectif(
                code_produit=code, mois=mois, annee=annee,
                distributor_id=current_distributor.id if current_distributor else None,
                objectif_tonne_vd=tonne_vd,
                objectif_packs_vd=packs_vd,
                objectif_packs_vd_tournee=packs_vd_t,
                objectif_tonne_vh=tonne_vh,
                objectif_packs_vh=packs_vh,
                objectif_packs_vh_tournee=packs_vh_t,
                created_by_id=current_user.id,
                updated_by_id=current_user.id,
                created_at=now, updated_at=now,
            ))

    session.commit()
    return {"ok": True, "saved": len(body)}


@router.get("")
def list_objectifs(
    mois: int = Query(..., ge=1, le=12),
    annee: int = Query(..., ge=2020),
    code_distributeur: Optional[str] = Query(None),
    distributor_id: Optional[int] = Query(None),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    current_distributor = Depends(get_current_distributor),
) -> Any:
    q = select(Objectif).where(
        Objectif.mois == mois,
        Objectif.annee == annee,
        (Objectif.objectif_tonne_vd.isnot(None)) |
        (Objectif.objectif_packs_vd.isnot(None)) |
        (Objectif.objectif_tonne_vh.isnot(None)) |
        (Objectif.objectif_packs_vh.isnot(None)),
    ).order_by(Objectif.code_produit)

    if current_user.role != UserRole.PLATFORM_ADMIN and current_distributor:
        q = q.where(Objectif.distributor_id == current_distributor.id)
    elif distributor_id:
        q = q.where(Objectif.distributor_id == distributor_id)
    elif code_distributeur:
        q = q.where(Objectif.code_distributeur == code_distributeur)

    rows = session.exec(q).all()

    user_ids = {obj.updated_by_id for obj in rows if obj.updated_by_id}
    users_map: dict = {}
    if user_ids:
        users = session.exec(select(User).where(User.id.in_(user_ids))).all()
        users_map = {u.id: u.full_name for u in users}

    return [
        {
            "code_produit": obj.code_produit,
            "nom_distributeur": obj.nom_distributeur,
            "objectif_tonne_vd": obj.objectif_tonne_vd,
            "objectif_tonne_vd_tournee": obj.objectif_tonne_vd_tournee,
            "objectif_packs_vd": obj.objectif_packs_vd,
            "objectif_packs_vd_tournee": obj.objectif_packs_vd_tournee,
            "objectif_tonne_vh": obj.objectif_tonne_vh,
            "objectif_tonne_vh_tournee": obj.objectif_tonne_vh_tournee,
            "objectif_packs_vh": obj.objectif_packs_vh,
            "objectif_packs_vh_tournee": obj.objectif_packs_vh_tournee,
            "updated_at": obj.updated_at.isoformat() if obj.updated_at else None,
            "updated_by": users_map.get(obj.updated_by_id) if obj.updated_by_id else None,
        }
        for obj in rows
    ]


@router.post("/parse-excel")
async def parse_excel(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> Any:
    import openpyxl
    # TODO: Produit model needs to be created for alias mapping
    dd_map = {}

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    result = []
    header = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = [str(c).strip().lower() if c else "" for c in row]
            has_code = header[0] == "code"
            continue
        if not row:
            continue
        if has_code:
            code = row[0]
            if not code or str(code).startswith("⚠"):
                continue
            code = dd_map.get(str(code).strip(), str(code).strip())
            tonne   = float(row[2]) if len(row) > 2 and row[2] is not None else None
            packs   = float(row[3]) if len(row) > 3 and row[3] is not None else None
            packs_t = float(row[4]) if len(row) > 4 and row[4] is not None else None
            result.append({"code_produit": code, "tonne": tonne, "packs": packs, "packs_tournee": packs_t})
        else:
            nom = row[0]
            if not nom:
                continue
            packs   = float(row[1]) if len(row) > 1 and row[1] is not None else None
            packs_t = float(row[2]) if len(row) > 2 and row[2] is not None else None
            result.append({"code_produit": None, "nom_produit": str(nom).strip(), "packs": packs, "packs_tournee": packs_t})
    return result


@router.post("/preview")
async def preview_objectifs(
    file: UploadFile = File(...),
    canal: str = Query(..., regex="^(VD|VH)$"),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
) -> Any:
    """Preview objectifs Excel file without saving"""

    content = await file.read()

    try:
        import openpyxl
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        except Exception:
            return {"error": "Format de fichier invalide. Veuillez utiliser un fichier Excel (.xlsx ou .xls)"}
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"error": "Fichier vide"}

        # Parse header
        header_raw = rows[0]
        header_lower = [str(c).strip().lower() if c else "" for c in header_raw]

        # Validate required columns
        code_idx = next((i for i, h in enumerate(header_lower) if 'code' in h), None)
        mois_idx = next((i for i, h in enumerate(header_lower) if 'mois' in h), None)
        tonne_idx = next((i for i, h in enumerate(header_lower) if 'tonne' in h), None)
        pack_idx  = next((i for i, h in enumerate(header_lower) if 'pack'  in h), None)

        missing = []
        if code_idx  is None: missing.append("Code")
        if mois_idx  is None: missing.append("Mois")
        if tonne_idx is None: missing.append("Tonne")
        if pack_idx  is None: missing.append("Pack")

        if missing:
            return {"error": f"Colonnes manquantes dans le fichier : {', '.join(missing)}. Vérifiez que le fichier est bien un fichier objectifs."}

        logger.info(f"Preview: header={header_raw}")

        # Normalised header labels (keep originals for display)
        headers = [str(c).strip() if c else "" for c in header_raw]

        # Extract mois/annee, count rows and collect all column values per product
        mois = None
        annee = None
        row_count = 0
        products = []

        for row_num, row in enumerate(rows[1:], 2):
            if not row or all(c is None for c in row):
                continue

            code = row[code_idx] if code_idx < len(row) else None
            if not code or str(code).startswith("⚠"):
                continue

            row_count += 1

            # Build a dict of all column values for this row
            row_data: dict = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                val = row[i] if i < len(row) else None
                if val is None:
                    row_data[h] = None
                elif hasattr(val, 'month') and hasattr(val, 'year'):
                    row_data[h] = f"{val.month:02d}/{val.year}"
                elif isinstance(val, float):
                    rounded = round(val, 3)
                    row_data[h] = int(rounded) if rounded == int(rounded) else rounded
                else:
                    row_data[h] = val
            products.append(row_data)

            # Extract mois/annee from first valid row
            if mois is None and mois_idx < len(row):
                mois_val = row[mois_idx]
                if mois_val:
                    try:
                        if hasattr(mois_val, 'month') and hasattr(mois_val, 'year'):
                            mois = mois_val.month
                            annee = mois_val.year
                        else:
                            mois_str = str(mois_val).strip()
                            if '/' in mois_str:
                                parts = mois_str.split('/')
                                if len(parts) == 2:
                                    mois = int(parts[0].strip())
                                    annee = int(parts[1].strip())
                            else:
                                logger.warning(f"Could not parse mois value: {mois_val} (type: {type(mois_val)})")
                    except (ValueError, AttributeError) as e:
                        logger.error(f"Failed to parse mois: {mois_val} - {e}")

        logger.info(f"Preview: extracted mois={mois}, annee={annee}, row_count={row_count}")

        if mois is None or annee is None:
            return {"error": f"Impossible d'extraire le mois/année du fichier (trouvé {row_count} lignes)"}

        return {"mois": mois, "annee": annee, "rowCount": row_count, "headers": headers, "products": products}

    except Exception as e:
        logger.error(f"Preview failed: {e}", exc_info=True)
        return {"error": f"Erreur lors de la lecture du fichier: {e}"}


@router.post("/upload")
async def upload_objectifs(
    file: UploadFile = File(...),
    canal: str = Query(..., regex="^(VD|VH)$"),
    distributor_id: Optional[int] = Query(None),
    route_count: int = Query(..., ge=1),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
    current_distributor = Depends(get_current_distributor),
) -> StreamingResponse:
    """Upload objectifs Excel file and save to database"""

    # Determine which distributor to use
    if distributor_id:
        # PLATFORM_ADMIN can upload for any distributor
        if current_user.role != UserRole.PLATFORM_ADMIN:
            raise HTTPException(status_code=403, detail="Only platform admins can upload for other distributors")
        target_distributor = session.exec(select(Distributor).where(Distributor.id == distributor_id)).first()
        if not target_distributor:
            raise HTTPException(status_code=404, detail="Distributor not found")
    else:
        target_distributor = current_distributor
        if not target_distributor:
            raise HTTPException(status_code=400, detail="No distributor specified")

    content = await file.read()
    filename = file.filename or ''

    def event(data: dict) -> str:
        return f"data: {json.dumps(data, ensure_ascii=False)}\n\n"

    def generate():
        logger.info(f"Objectifs upload démarré : {filename} ({canal}) par {current_user.full_name} ({len(content) // 1024} KB)")
        yield event({"progress": 5, "message": "Lecture du fichier..."})

        try:
            import openpyxl
            try:
                wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            except Exception:
                yield event({"error": "Format de fichier invalide. Veuillez utiliser un fichier Excel (.xlsx ou .xls)"})
                return
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                yield event({"error": "Fichier vide"})
                return

            # Parse and validate header
            header_lower = [str(c).strip().lower() if c else "" for c in rows[0]]

            code_idx  = next((i for i, h in enumerate(header_lower) if 'code'  in h), None)
            mois_idx  = next((i for i, h in enumerate(header_lower) if 'mois'  in h), None)
            tonne_idx = next((i for i, h in enumerate(header_lower) if 'tonne' in h), None)
            pack_idx  = next((i for i, h in enumerate(header_lower) if 'pack'  in h), None)

            missing = []
            if code_idx  is None: missing.append("Code")
            if mois_idx  is None: missing.append("Mois")
            if tonne_idx is None: missing.append("Tonne")
            if pack_idx  is None: missing.append("Pack")

            if missing:
                yield event({"error": f"Colonnes manquantes : {', '.join(missing)}. Vérifiez que le fichier est bien un fichier objectifs."})
                return

            # Parse data rows
            data = []
            mois = None
            annee = None

            for row_num, row in enumerate(rows[1:], 2):
                if not row or all(c is None for c in row):
                    continue

                code = row[code_idx] if code_idx < len(row) else None
                if not code or str(code).startswith("⚠"):
                    continue

                code = str(code).strip()

                # Extract mois/annee from first row (should be same for all)
                if mois is None:
                    mois_val = row[mois_idx] if mois_idx < len(row) else None
                    if mois_val:
                        try:
                            # openpyxl returns datetime or date objects for date cells
                            if hasattr(mois_val, 'month') and hasattr(mois_val, 'year'):
                                mois = mois_val.month
                                annee = mois_val.year
                            else:
                                parts = str(mois_val).split('/')
                                if len(parts) == 2:
                                    mois = int(parts[0].strip())
                                    annee = int(parts[1].strip())
                        except (ValueError, AttributeError):
                            yield event({"error": f"Format Mois invalide à la ligne {row_num}: {mois_val}"})
                            return

                tonne_raw = float(row[tonne_idx]) if tonne_idx < len(row) and row[tonne_idx] is not None else None
                packs_raw = float(row[pack_idx]) if pack_idx < len(row) and row[pack_idx] is not None else None

                # Per-route calculations
                tonne_tournee = tonne_raw / route_count if tonne_raw is not None else None
                packs_tournee = packs_raw / route_count if packs_raw is not None else None

                # Transform based on canal
                item = {"code_produit": code}
                if canal == "VD":
                    item["objectif_tonne_vd"] = tonne_raw
                    item["objectif_tonne_vd_tournee"] = tonne_tournee
                    item["objectif_packs_vd"] = packs_raw
                    item["objectif_packs_vd_tournee"] = packs_tournee
                else:  # VH
                    item["objectif_tonne_vh"] = tonne_raw
                    item["objectif_tonne_vh_tournee"] = tonne_tournee
                    item["objectif_packs_vh"] = packs_raw
                    item["objectif_packs_vh_tournee"] = packs_tournee

                data.append(item)

            if mois is None or annee is None:
                yield event({"error": "Impossible d'extraire le mois/année du fichier"})
                return

            if not data:
                yield event({"error": "Aucune ligne valide trouvée"})
                return

            total_rows = len(data)
            yield event({
                "progress": 15,
                "message": f"{total_rows:,} lignes trouvées ({mois:02d}/{annee})...",
                "file_info": {"total_rows": total_rows, "date_min": f"{mois:02d}/{annee}", "date_max": f"{mois:02d}/{annee}"},
            })

            yield event({"progress": 30, "message": "Suppression des anciens objectifs..."})

            now = datetime.now(timezone.utc)

            # Load all existing rows for this distributor/mois/annee
            existing = session.exec(
                select(Objectif).where(
                    Objectif.mois == mois,
                    Objectif.annee == annee,
                    Objectif.distributor_id == target_distributor.id,
                )
            ).all()
            obj_map = {o.code_produit: o for o in existing}

            # Step 1 — clear the canal being replaced on every existing row
            for obj in existing:
                if canal == "VD":
                    obj.objectif_tonne_vd         = None
                    obj.objectif_tonne_vd_tournee = None
                    obj.objectif_packs_vd         = None
                    obj.objectif_packs_vd_tournee = None
                else:
                    obj.objectif_tonne_vh         = None
                    obj.objectif_tonne_vh_tournee = None
                    obj.objectif_packs_vh         = None
                    obj.objectif_packs_vh_tournee = None

            yield event({"progress": 50, "message": "Sauvegarde en base de données..."})

            # Step 2 — upsert uploaded rows
            saved_count = 0
            for item in data:
                code = item.get("code_produit")
                if not code:
                    continue

                obj = obj_map.get(code)
                if obj:
                    obj.code_distributeur = target_distributor.code
                    obj.nom_distributeur  = target_distributor.nom
                    if canal == "VD":
                        obj.objectif_tonne_vd          = item.get("objectif_tonne_vd")
                        obj.objectif_tonne_vd_tournee  = item.get("objectif_tonne_vd_tournee")
                        obj.objectif_packs_vd          = item.get("objectif_packs_vd")
                        obj.objectif_packs_vd_tournee  = item.get("objectif_packs_vd_tournee")
                    else:
                        obj.objectif_tonne_vh          = item.get("objectif_tonne_vh")
                        obj.objectif_tonne_vh_tournee  = item.get("objectif_tonne_vh_tournee")
                        obj.objectif_packs_vh          = item.get("objectif_packs_vh")
                        obj.objectif_packs_vh_tournee  = item.get("objectif_packs_vh_tournee")
                    obj.updated_by_id = current_user.id
                    obj.updated_at    = now
                else:
                    session.add(Objectif(
                        code_produit=code,
                        mois=mois,
                        annee=annee,
                        distributor_id=target_distributor.id,
                        code_distributeur=target_distributor.code,
                        nom_distributeur=target_distributor.nom,
                        objectif_tonne_vd=item.get("objectif_tonne_vd"),
                        objectif_tonne_vd_tournee=item.get("objectif_tonne_vd_tournee"),
                        objectif_packs_vd=item.get("objectif_packs_vd"),
                        objectif_packs_vd_tournee=item.get("objectif_packs_vd_tournee"),
                        objectif_tonne_vh=item.get("objectif_tonne_vh"),
                        objectif_tonne_vh_tournee=item.get("objectif_tonne_vh_tournee"),
                        objectif_packs_vh=item.get("objectif_packs_vh"),
                        objectif_packs_vh_tournee=item.get("objectif_packs_vh_tournee"),
                        created_by_id=current_user.id,
                        updated_by_id=current_user.id,
                        created_at=now,
                        updated_at=now,
                    ))
                saved_count += 1

            # Step 3 — delete rows where both canals are now fully empty
            for obj in existing:
                has_vd = any([obj.objectif_tonne_vd, obj.objectif_packs_vd])
                has_vh = any([obj.objectif_tonne_vh, obj.objectif_packs_vh])
                if not has_vd and not has_vh:
                    session.delete(obj)

            session.commit()

            logger.info(f"Objectifs upload terminé : {saved_count:,} lignes sauvegardées ({canal}) par {current_user.full_name}")
            yield event({
                "progress": 95,
                "message": f"{saved_count:,} objectifs sauvegardés..."
            })

            yield event({
                "progress": 100,
                "done": True,
                "message": f"{saved_count:,} objectifs {canal} importés avec succès"
            })

        except Exception as e:
            logger.error(f"Objectifs upload échoué : {filename} — {e}")
            yield event({"error": f"Erreur lors de l'import : {e}"})

    return StreamingResponse(generate(), media_type="text/event-stream")
