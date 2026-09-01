from typing import Any, List, Optional
from collections import defaultdict
from datetime import datetime, timezone
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Body, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, case as sa_case, text
from sqlmodel import Session, select

from app.api.deps import get_current_user, get_current_distributor
from app.core.product_codes import build_produit_map, load_obj_to_vente_map, obj_val, remap_by_vente_code, remap_dict
from app.core.uom_conversion import PRODUITS_JOIN, qty_expr
from app.database import get_session
from app.models.user import User, UserRole
from app.models.vente import Vente
from app.models.objectif import Objectif
from app.models.produit import Produit

router = APIRouter()


@router.get("/periodes", response_model=List[str])
def prevendeur_periodes(
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
) -> Any:
    if not current_user.employe_code:
        return []
    rows = session.exec(
        select(Vente.annee_mois)
        .distinct()
        .where(Vente.code_fdv == current_user.employe_code)
        .order_by(Vente.annee_mois.desc())
    ).all()
    return list(rows)


@router.get("/facturation")
def prevendeur_facturation(
    annee_mois: str = Query(...),
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
) -> Any:
    if not current_user.employe_code:
        return {"fdv_nom": "", "periode": annee_mois, "routes": [], "products": [], "products_meta": {}, "total_clients": 0}

    rows = session.exec(
        select(Vente)
        .where(Vente.code_fdv == current_user.employe_code)
        .where(Vente.annee_mois == annee_mois)
        .where(Vente.famille.ilike('sucre') | Vente.famille.ilike('huile'))
    ).all()

    # Exclude explicit BackOffice rows
    rows = [r for r in rows if r.source != 'BackOffice']

    if not rows:
        return {"fdv_nom": current_user.full_name, "periode": annee_mois, "routes": [], "products": [], "products_meta": {}, "total_clients": 0}

    # Use Vente data directly since Produit model was removed
    code_to_label = {r.code_produit: (r.description_produit or f"Code {r.code_produit}") for r in rows if r.code_produit}

    # Filter by available vente data (facturable flag no longer available)
    # rows = [r for r in rows if r.code_produit and r.code_produit in code_to_label]

    def display_label(r: Vente) -> str:
        if r.code_produit and r.code_produit in code_to_label and code_to_label[r.code_produit]:
            return code_to_label[r.code_produit]
        return r.description_produit or ''

    products = sorted({display_label(r) for r in rows if r.description_produit})

    # Build meta map from Vente data only (Produit model removed)
    label_meta: dict = {}
    for r in rows:
        label = display_label(r)
        if label not in label_meta:
            famille = (r.famille or '').lower()
            label_meta[label] = {"uom_vente": r.uom_vente, "colisage": None, "famille": famille, "prix": None}
    products_meta = {p: label_meta.get(p, {"uom_vente": None, "colisage": None, "famille": None, "prix": None}) for p in products}

    # Aggregate: client -> date_label -> product -> qty
    agg: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    client_meta: dict = {}

    for r in rows:
        if not r.nom_client or not r.description_produit or not r.date_commande:
            continue
        label = r.date_commande.strftime('%d/%m/%y')
        agg[r.nom_client][label][display_label(r)] += r.qte_facturee or 0
        if r.nom_client not in client_meta:
            client_meta[r.nom_client] = {
                'code_client': r.code_client,
                'route': r.route or 'Sans route',
            }

    all_dates = sorted({r.date_commande for r in rows if r.date_commande})
    date_labels = [d.strftime('%d/%m/%y') for d in all_dates]

    # Group clients by route
    route_clients: dict = defaultdict(list)
    for nom, meta in sorted(client_meta.items()):
        route_clients[meta['route']].append(nom)

    fdv_nom = rows[0].nom_fdv if rows else current_user.full_name

    # Client metadata
    nom_sodichn_map = {}

    routes_out = []
    for route in sorted(route_clients.keys()):
        clients_out = []
        for nom in route_clients[route]:
            client_dates = [l for l in date_labels if any(agg[nom][l][p] for p in products)]
            if not client_dates:
                continue
            semaines = {
                l: {p: (agg[nom][l][p] if agg[nom][l][p] else None) for p in products}
                for l in client_dates
            }
            totaux = {p: (sum(agg[nom][l][p] for l in client_dates) or None) for p in products}
            code = client_meta[nom]['code_client']
            clients_out.append({
                "nom_client": nom,
                "code_client": code,
                "nom_sodichn": nom_sodichn_map.get(code),
                "derniere_visite": client_dates[-1],
                "weeks": client_dates,
                "semaines": semaines,
                "totaux": totaux,
            })
        if clients_out:
            routes_out.append({"route": route, "clients": clients_out})

    return {
        "fdv_nom": fdv_nom,
        "periode": annee_mois,
        "products": products,
        "products_meta": products_meta,
        "total_clients": sum(len(r["clients"]) for r in routes_out),
        "routes": routes_out,
    }


@router.get("/admin/stats")
def prevendeur_admin_stats(
    current_user: User = Depends(get_current_user),
    current_distributor = Depends(get_current_distributor),
    session: Session = Depends(get_session),
) -> Any:
    q = select(User).where(User.role == UserRole.PREVENDEUR).where(User.is_active == True)
    if current_distributor:
        q = q.where(User.distributor_id == current_distributor.id)
    prevendeurs = session.exec(q.order_by(User.full_name)).all()

    fdv_codes = [pv.employe_code for pv in prevendeurs if pv.employe_code]
    if not fdv_codes:
        return []

    # 1 query: all distinct (code_fdv, code_client, nom_client) across all prevendeurs
    vente_rows = session.exec(
        select(Vente.code_fdv, Vente.code_client, Vente.nom_client)
        .distinct()
        .where(Vente.code_fdv.in_(fdv_codes), Vente.code_client.isnot(None))
    ).all()

    fdv_clients: dict = defaultdict(list)   # code_fdv -> [(code_client, nom_client)]
    all_client_codes: set = set()
    for code_fdv, code_client, nom_client in vente_rows:
        fdv_clients[code_fdv].append((code_client, nom_client))
        all_client_codes.add(code_client)

    # 1 query: last sale date per prevendeur
    last_sale_rows = session.exec(
        select(Vente.code_fdv, func.max(Vente.date_commande))
        .where(Vente.code_fdv.in_(fdv_codes))
        .group_by(Vente.code_fdv)
    ).all()
    last_sale_map = {code_fdv: last_sale for code_fdv, last_sale in last_sale_rows}

    # Client metadata
    sodichn_map: dict = {}   # customer_no -> {nom_sodichn, updated_at}

    today = datetime.now(timezone.utc).date()

    result = []
    for pv in prevendeurs:
        if not pv.employe_code:
            continue
        clients = fdv_clients.get(pv.employe_code, [])
        total_clients = len(clients)

        # Per-client stats
        updated_today = 0
        last_sodichn_date = None
        matched = 0
        clients_detail = []

        for code, nom in sorted(clients, key=lambda x: x[1] or ''):
            info = sodichn_map.get(code, {})
            nom_sodichn = info.get("nom_sodichn")
            updated_at = info.get("updated_at")
            if nom_sodichn:
                matched += 1
            if updated_at and nom_sodichn:
                updated_date = updated_at.date() if hasattr(updated_at, 'date') else updated_at
                if updated_date == today:
                    updated_today += 1
                if last_sodichn_date is None or updated_date > last_sodichn_date:
                    last_sodichn_date = updated_date
            clients_detail.append({
                "code_client": code,
                "nom_client": nom,
                "nom_sodichn": nom_sodichn,
                "updated_at": updated_at.strftime('%Y-%m-%d') if updated_at else None,
            })

        # Count updated on last active day
        updated_on_last_day = 0
        if last_sodichn_date:
            updated_on_last_day = sum(
                1 for c in clients_detail
                if c["updated_at"] and c["updated_at"][:10] == last_sodichn_date.strftime('%Y-%m-%d')
            )

        last_sale = last_sale_map.get(pv.employe_code)
        result.append({
            "id": pv.id,
            "full_name": pv.full_name,
            "employe_code": pv.employe_code,
            "total_clients": total_clients,
            "clients_with_sodichn": matched,
            "remaining": total_clients - matched,
            "completion_pct": round(matched / total_clients * 100) if total_clients > 0 else 0,
            "updated_today": updated_today,
            "last_sodichn_date": last_sodichn_date.strftime('%Y-%m-%d') if last_sodichn_date else None,
            "updated_on_last_day": updated_on_last_day,
            "last_activity": last_sale.strftime('%Y-%m-%d') if last_sale else None,
            "clients": clients_detail,
        })

    return result


@router.get("/admin/stats/export")
def export_clients_excel(
    current_user: User = Depends(get_current_user),
    current_distributor = Depends(get_current_distributor),
    session: Session = Depends(get_session),
):
    q = select(User).where(User.role == UserRole.PREVENDEUR).where(User.is_active == True)
    if current_distributor:
        q = q.where(User.distributor_id == current_distributor.id)
    prevendeurs = session.exec(q.order_by(User.full_name)).all()

    fdv_codes = [pv.employe_code for pv in prevendeurs if pv.employe_code]
    if not fdv_codes:
        return StreamingResponse(BytesIO(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    vente_rows = session.exec(
        select(Vente.code_fdv, Vente.code_client, Vente.nom_client)
        .distinct()
        .where(Vente.code_fdv.in_(fdv_codes), Vente.code_client.isnot(None))
    ).all()

    fdv_clients: dict = defaultdict(list)
    all_client_codes: set = set()
    for code_fdv, code_client, nom_client in vente_rows:
        fdv_clients[code_fdv].append((code_client, nom_client))
        all_client_codes.add(code_client)

    sodichn_map: dict = {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients RC"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="2563EB")
    fdv_font = Font(bold=True, size=10)
    fdv_fill = PatternFill("solid", fgColor="DBEAFE")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    headers = ["Prévendeur", "Code FDV", "Code Client", "Nom Client", "Nom RC (Sodichn)", "Statut", "Mis à jour le"]
    col_widths = [28, 14, 14, 32, 32, 12, 16]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 22

    row_num = 2
    today = datetime.now(timezone.utc).date()

    for pv in prevendeurs:
        if not pv.employe_code:
            continue
        clients = fdv_clients.get(pv.employe_code, [])
        for code, nom in sorted(clients, key=lambda x: x[1] or ''):
            info = sodichn_map.get(code, {})
            nom_sodichn = info.get("nom_sodichn") or ""
            updated_at = info.get("updated_at")
            statut = "✓ Complété" if nom_sodichn else "— Vide"
            updated_str = updated_at.strftime('%Y-%m-%d') if updated_at else ""

            row_data = [pv.full_name, pv.employe_code, code, nom, nom_sodichn, statut, updated_str]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = border
                cell.font = Font(size=9)
                if col == 6:
                    cell.alignment = center
                    if nom_sodichn:
                        cell.font = Font(size=9, color="16A34A", bold=True)
                    else:
                        cell.font = Font(size=9, color="9CA3AF")

            row_num += 1

    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"clients_rc_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/admin/drilldown")
def prevendeur_admin_drilldown(
    annee_mois: str = Query(...),
    code_fdv: Optional[str] = Query(None),
    canal: Optional[str] = Query(None),   # "VD" or "VH"
    nom_distributeur: Optional[str] = Query(None),
    unite: str = Query("tonnes"),
    current_user: User = Depends(get_current_user),
    current_distributor = Depends(get_current_distributor),
    session: Session = Depends(get_session),
) -> Any:
    q_periods = select(Vente.annee_mois).distinct().order_by(Vente.annee_mois.desc())
    if current_distributor:
        q_periods = q_periods.where((Vente.distributor_id == current_distributor.id) | (Vente.distributor_id == None))
    all_periods = session.exec(q_periods).all()
    all_periods_list = list(all_periods)

    # Prevendeurs list with their totals for the current period (always unfiltered)
    q_pv = select(User).where(User.role == UserRole.PREVENDEUR).where(User.is_active == True)
    if current_distributor:
        q_pv = q_pv.where(User.distributor_id == current_distributor.id)
    prevendeurs_db = session.exec(q_pv.order_by(User.full_name)).all()
    fdv_name_map = {p.employe_code: p.full_name for p in prevendeurs_db if p.employe_code}

    # Quantity expression: tonnes (via produit weight) or raw packs
    if unite == "tonnes":
        _norm = sa_case(
            (Produit.poids_unite_vente.isnot(None),
             Vente.qte_facturee * Produit.poids_unite_vente / 1000),
            else_=0,
        )
        _rq = lambda v: round(v, 2)
    else:
        _norm = Vente.qte_facturee
        _rq = round

    def _produit_join(q):
        """LEFT JOIN Vente with Produit on code_produit."""
        return q.join(Produit, Vente.code_produit == Produit.code_produit, isouter=True)

    def _tonnes_join(q):
        """Add Produit JOIN only when converting to tonnes."""
        return _produit_join(q) if unite == "tonnes" else q

    fdv_totals_q = _tonnes_join(
        select(Vente.code_fdv, func.sum(_norm))
        .where(Vente.annee_mois == annee_mois)
        .where(Vente.statut_commande == 'Facturé')
        .where(Vente.code_fdv.isnot(None))
        .group_by(Vente.code_fdv)
    )
    if current_distributor:
        fdv_totals_q = fdv_totals_q.where((Vente.distributor_id == current_distributor.id) | (Vente.distributor_id == None))
    if canal:
        fdv_totals_q = fdv_totals_q.where(Vente.canal == canal)
    if nom_distributeur:
        fdv_totals_q = fdv_totals_q.where(Vente.nom_distributeur == nom_distributeur)
    fdv_totals = {row[0]: _rq(row[1] or 0) for row in session.exec(fdv_totals_q).all()}

    canal_upper = canal.upper() if canal else None

    # Previous period — use the previous available period in the data (not necessarily month-1)
    try:
        cur_idx = all_periods_list.index(annee_mois)
    except ValueError:
        cur_idx = 0
    prev_periode = all_periods_list[cur_idx + 1] if cur_idx + 1 < len(all_periods_list) else None
    trend_periods = list(reversed(all_periods_list[cur_idx: cur_idx + 6]))  # chronological

    def fetch_rows(periode: str) -> list:
        q = _produit_join(select(
            Vente.date_commande,
            Vente.famille,
            Vente.sous_famille,
            Vente.code_produit,
            Vente.description_produit,
            _norm.label('qty_norm'),
            Vente.code_fdv,
            Vente.nom_fdv,
            Vente.source,
            Produit.famille.label('produit_famille'),
            Produit.sous_famille.label('produit_sous_famille'),
        )).where(
            Vente.annee_mois == periode, Vente.statut_commande == 'Facturé'
        )
        if current_distributor:
            q = q.where((Vente.distributor_id == current_distributor.id) | (Vente.distributor_id == None))
        if code_fdv:
            q = q.where(Vente.code_fdv == code_fdv)
        if canal:
            q = q.where(Vente.canal == canal)
        if nom_distributeur:
            q = q.where(Vente.nom_distributeur == nom_distributeur)
        return list(session.exec(q).all())

    rows = fetch_rows(annee_mois)
    prev_rows = fetch_rows(prev_periode) if prev_periode else []

    period_totals: dict = defaultdict(float)
    if trend_periods:
        q6 = _tonnes_join(
            select(Vente.annee_mois, func.sum(_norm))
            .where(Vente.annee_mois.in_(trend_periods))
            .where(Vente.statut_commande == 'Facturé')
        )
        if current_distributor:
            q6 = q6.where((Vente.distributor_id == current_distributor.id) | (Vente.distributor_id == None))
        if code_fdv:
            q6 = q6.where(Vente.code_fdv == code_fdv)
        if canal:
            q6 = q6.where(Vente.canal == canal)
        if nom_distributeur:
            q6 = q6.where(Vente.nom_distributeur == nom_distributeur)
        q6 = q6.group_by(Vente.annee_mois)
        for periode, total in session.exec(q6).all():
            period_totals[periode] = total or 0

    trend_6m = [_rq(period_totals.get(p, 0)) for p in trend_periods]
    trend_6m_labels = trend_periods

    # Product label resolution — Produit model removed, use description_produit from Vente
    codes = {r.code_produit for r in rows if r.code_produit}
    code_to_produit = {}  # No Produit model available

    def _norm_fam(val) -> str:
        return (val or '').strip().lower()

    def week_idx(d) -> int:
        if d is None:
            return 0
        day = d.day
        if day <= 7: return 0
        if day <= 14: return 1
        if day <= 21: return 2
        return 3

    def product_label(r: Vente) -> str:
        return r.description_produit or "Autre"

    # famille -> sous_famille -> produit -> [w0,w1,w2,w3]
    hier: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: [0.0] * 4)))
    # famille -> code_fdv -> total
    fdv_by_famille: dict = defaultdict(lambda: defaultdict(float))
    prev_famille_total: dict = defaultdict(float)
    # famille -> sf -> produit -> code_fdv -> total (for per-product FDV panel)
    fdv_by_sf_prod: dict = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(float))))
    # product label -> code_produit (first seen)
    prod_label_to_code: dict = {}

    for r in rows:
        if not r.date_commande or not r.qty_norm:
            continue
        famille = _norm_fam(r.produit_famille or r.famille)
        sf = (r.produit_sous_famille or r.sous_famille or "Autres").strip()
        prod = product_label(r)
        w = week_idx(r.date_commande)
        hier[famille][sf][prod][w] += r.qty_norm
        if r.code_produit and prod not in prod_label_to_code:
            prod_label_to_code[prod] = r.code_produit
        if r.code_fdv:
            fdv_by_famille[famille][r.code_fdv] += r.qty_norm
            fdv_by_sf_prod[famille][sf][prod][r.code_fdv] += r.qty_norm
        # Supplement fdv name map from row data
        if r.code_fdv and r.nom_fdv and r.code_fdv not in fdv_name_map:
            fdv_name_map[r.code_fdv] = r.nom_fdv

    for r in prev_rows:
        if not r.qty_norm:
            continue
        famille = _norm_fam(r.produit_famille or r.famille)
        prev_famille_total[famille] += r.qty_norm

    # Objective aggregation for this period
    annee_int, mois_int = int(annee_mois.split('-')[0]), int(annee_mois.split('-')[1])

    # Per-product objectives (total + per-tournée)
    obj_q = (
        select(
            Objectif.code_produit,
            Objectif.objectif_packs_vd, Objectif.objectif_packs_vh,
            Objectif.objectif_packs_vd_tournee, Objectif.objectif_packs_vh_tournee,
            Objectif.objectif_tonne_vd, Objectif.objectif_tonne_vh,
        )
        .where(Objectif.mois == mois_int, Objectif.annee == annee_int)
    )
    if current_distributor:
        obj_q = obj_q.where(Objectif.distributor_id == current_distributor.id)
    obj_prod_rows = session.exec(obj_q).all()

    obj_by_prod       = {r.code_produit: obj_val(r, 'objectif_packs', canal, '_tournee') for r in obj_prod_rows}
    obj_by_prod_total = {r.code_produit: obj_val(r, 'objectif_packs', canal) for r in obj_prod_rows}
    obj_tonne_by_prod = {r.code_produit: obj_val(r, 'objectif_tonne', canal) for r in obj_prod_rows}

    # Remap objective codes (may be code_dd) → vente codes so sales lookups match
    obj_to_vente = load_obj_to_vente_map(session)
    obj_by_prod       = remap_dict(obj_by_prod, obj_to_vente)
    obj_by_prod_total = remap_dict(obj_by_prod_total, obj_to_vente)
    obj_tonne_by_prod = remap_dict(obj_tonne_by_prod, obj_to_vente)

    # Canonical famille objectives via Produit mapping (includes zero-sale products)
    _produit_map = build_produit_map({r.code_produit for r in obj_prod_rows}, session)

    obj_tonne_by_famille: dict = defaultdict(float)
    obj_packs_by_famille: dict = defaultdict(float)
    obj_tonne_by_sf: dict = defaultdict(float)
    obj_packs_by_sf: dict = defaultdict(float)
    for r in obj_prod_rows:
        p = _produit_map.get(r.code_produit)
        fam = _norm_fam(p.famille) if p else ''
        sf = (p.sous_famille or 'Autres').strip() if p else 'Autres'
        fam_sf_key = f"{fam}||{sf}"
        tonne = obj_val(r, 'objectif_tonne', canal)
        packs = obj_val(r, 'objectif_packs', canal)
        obj_tonne_by_famille[fam] += tonne
        obj_packs_by_famille[fam] += packs
        obj_tonne_by_sf[fam_sf_key] += tonne
        obj_packs_by_sf[fam_sf_key] += packs

    # When a single FDV is selected, display per-route targets instead of global totals
    if code_fdv:
        obj_by_prod_total = obj_by_prod

    # Per-route total: sum of all tournée targets (derived from obj_by_prod, no extra query)
    objectif_per_route = round(sum(v for v in obj_by_prod.values() if v))

    # Per-fdv per-product sales — queried without code_fdv filter so all pills stay accurate
    _fdv_prod_q = _tonnes_join(
        select(Vente.code_fdv, Vente.code_produit, func.sum(_norm))
        .where(Vente.annee_mois == annee_mois, Vente.statut_commande == 'Facturé', Vente.code_fdv.isnot(None), Vente.code_produit.isnot(None))
    )
    if current_distributor:
        _fdv_prod_q = _fdv_prod_q.where((Vente.distributor_id == current_distributor.id) | (Vente.distributor_id == None))
    if canal:
        _fdv_prod_q = _fdv_prod_q.where(Vente.canal == canal)
    _fdv_prod_q = _fdv_prod_q.group_by(Vente.code_fdv, Vente.code_produit)

    fdv_prod_totals: dict = defaultdict(lambda: defaultdict(float))
    for _fdv, _prod, _tot in session.exec(_fdv_prod_q).all():
        fdv_prod_totals[_fdv][_prod] = _tot or 0.0

    def compute_achievement_pct(code_fdv: str):
        """Simple average of per-product achievement rates for this prevendeur."""
        rates = []
        prod_sales = fdv_prod_totals.get(code_fdv, {})
        for code_prod, obj in obj_by_prod.items():
            if not obj:
                continue
            sales = prod_sales.get(code_prod, 0.0)
            rates.append(min(sales / obj * 100, 100.0))
        if not rates:
            return None
        return round(sum(rates) / len(rates))

    prevendeurs_out = sorted(
        [
            {
                "code": code,
                "nom": fdv_name_map.get(code, code),
                "total": total,
                "achievement_pct": compute_achievement_pct(code),
            }
            for code, total in fdv_totals.items()
            if not canal_upper or canal_upper in code.upper()
        ],
        key=lambda x: x["nom"],
    )

    # CA (chiffre d'affaires) per famille — SUM(qte_facturee * prix_unitaire)
    _ca_famille = func.coalesce(Produit.famille, Vente.famille)

    def _ca_by_periode(periode: str) -> dict:
        q = _produit_join(
            select(_ca_famille, func.sum(Vente.qte_facturee * Vente.prix_unitaire))
            .where(Vente.annee_mois == periode, Vente.statut_commande == 'Facturé')
            .where(Vente.prix_unitaire.isnot(None))
        ).group_by(_ca_famille)
        if current_distributor:
            q = q.where((Vente.distributor_id == current_distributor.id) | (Vente.distributor_id == None))
        if code_fdv:
            q = q.where(Vente.code_fdv == code_fdv)
        if canal:
            q = q.where(Vente.canal == canal)
        return {_norm_fam(r[0]): round(r[1] or 0) for r in session.exec(q).all()}

    ca_by_famille = _ca_by_periode(annee_mois)
    ca_prev_by_famille = _ca_by_periode(prev_periode) if prev_periode else {}

    familles_out = []
    for famille, sf_map in sorted(hier.items()):
        f_weeks = [0.0] * 4
        sfs_out = []
        for sf, prod_map in sf_map.items():
            sf_weeks = [0.0] * 4
            prods_out = []
            for prod, wks in prod_map.items():
                for i in range(4):
                    sf_weeks[i] += wks[i]
                prod_top_fdv = sorted(
                    [{"code": c, "nom": fdv_name_map.get(c, c), "total": _rq(t)}
                     for c, t in fdv_by_sf_prod[famille][sf][prod].items()],
                    key=lambda x: -x["total"]
                )
                prod_code = prod_label_to_code.get(prod)
                prod_obj_t = obj_by_prod.get(prod_code) if prod_code else None
                prod_obj   = obj_by_prod_total.get(prod_code) if prod_code else None
                prod_tonne = obj_tonne_by_prod.get(prod_code) if prod_code else None
                prods_out.append({
                    "nom": prod,
                    "total": _rq(sum(wks)),
                    "weeks": [_rq(v) for v in wks],
                    "top_fdv": prod_top_fdv,
                    "objectif_packs": round(prod_obj) if prod_obj else None,
                    "objectif_packs_tournee": round(prod_obj_t) if prod_obj_t else None,
                    "objectif_tonne": round(prod_tonne, 3) if prod_tonne else None,
                })
            for i in range(4):
                f_weeks[i] += sf_weeks[i]
            sf_key = f"{famille}||{sf}"
            sf_tonne_canon = obj_tonne_by_sf.get(sf_key)
            sf_packs_canon = obj_packs_by_sf.get(sf_key)
            sfs_out.append({
                "nom": sf,
                "total": _rq(sum(sf_weeks)),
                "weeks": [_rq(v) for v in sf_weeks],
                "produits": sorted(prods_out, key=lambda x: -x["total"]),
                "objectif_packs": round(sf_packs_canon) if sf_packs_canon else None,
                "objectif_tonne": round(sf_tonne_canon, 3) if sf_tonne_canon else None,
            })

        f_total = _rq(sum(f_weeks))
        prev_total = _rq(prev_famille_total.get(famille, 0))
        delta_pct = round((f_total - prev_total) / prev_total * 100) if prev_total > 0 else None

        top_fdv = sorted(
            [{"code": c, "nom": fdv_name_map.get(c, c), "total": _rq(t)} for c, t in fdv_by_famille[famille].items()],
            key=lambda x: -x["total"]
        )

        # Family objective from canonical Produit mapping (includes zero-sale products)
        f_tonne_canon = obj_tonne_by_famille.get(famille)
        f_packs_canon = obj_packs_by_famille.get(famille)
        f_obj = round(f_packs_canon) if f_packs_canon else None
        f_tonne = round(f_tonne_canon, 3) if f_tonne_canon else None

        familles_out.append({
            "nom": famille,
            "total": f_total,
            "total_prev": prev_total,
            "delta_pct": delta_pct,
            "weeks": [_rq(v) for v in f_weeks],
            "sous_familles": sorted(sfs_out, key=lambda x: -x["total"]),
            "top_fdv": top_fdv,
            "objectif_packs": f_obj,
            "objectif_tonne": f_tonne,
            "ca": ca_by_famille.get(famille) or None,
            "ca_prev": ca_prev_by_famille.get(famille) or None,
        })

    # Global objective totals — summed directly from the objectifs table
    global_obj_tonne = sum(obj_val(r, 'objectif_tonne', canal) for r in obj_prod_rows)
    global_obj_packs = sum(obj_val(r, 'objectif_packs', canal) for r in obj_prod_rows)

    # Per-route objective in tonnes (derived from packs ratio)
    if objectif_per_route and global_obj_packs and global_obj_tonne:
        obj_tonne_per_route = round(objectif_per_route * global_obj_tonne / global_obj_packs, 3)
    else:
        obj_tonne_per_route = None

    return {
        "periode": annee_mois,
        "periodes": list(all_periods),
        "prevendeurs": prevendeurs_out,
        "trend_6m": trend_6m,
        "trend_6m_labels": trend_6m_labels,
        "familles": sorted(familles_out, key=lambda x: -x["total"]),
        "objectif_packs_per_route": objectif_per_route or None,
        "objectif_tonne_per_route": obj_tonne_per_route,
        "global_objectif_tonne": round(global_obj_tonne, 3) if global_obj_tonne else None,
        "global_objectif_packs": round(global_obj_packs) if global_obj_packs else None,
        "global_ca": round(sum(ca_by_famille.values())) if ca_by_famille else None,
    }


@router.get("/objectifs")
def prevendeur_objectifs(
    annee_mois: str = Query(...),
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
) -> Any:
    if not current_user.employe_code:
        return []

    code_fdv = current_user.employe_code
    annee_int, mois_int = int(annee_mois.split('-')[0]), int(annee_mois.split('-')[1])

    code_upper = code_fdv.upper()
    use_vd = 'VH' not in code_upper  # default to VD unless explicitly VH

    # Actual sales per product — with name fallback from Vente
    sales_rows = session.exec(
        select(Vente.code_produit, Vente.description_produit, func.sum(Vente.qte_facturee))
        .where(Vente.code_fdv == code_fdv)
        .where(Vente.annee_mois == annee_mois)
        .where(Vente.source != 'BackOffice')
        .group_by(Vente.code_produit, Vente.description_produit)
    ).all()

    sales_by_code: dict = {}
    desc_by_code: dict = {}
    for code, desc, qty in sales_rows:
        if not code:
            continue
        sales_by_code[code] = sales_by_code.get(code, 0) + (qty or 0)
        if desc and code not in desc_by_code:
            desc_by_code[code] = desc

    # Objectives for this period
    obj_rows = session.exec(
        select(Objectif)
        .where(Objectif.mois == mois_int, Objectif.annee == annee_int)
    ).all()

    raw_obj: dict = {}
    for obj in obj_rows:
        if not obj.code_produit:
            continue
        objectif = obj.objectif_packs_vd_tournee if use_vd else obj.objectif_packs_vh_tournee
        raw_obj[obj.code_produit] = objectif or 0

    # Remap objective codes (may be code_dd) → vente codes
    obj_by_code = remap_by_vente_code(raw_obj, session)

    all_codes = set(sales_by_code.keys()) | set(obj_by_code.keys())
    code_to_produit = {}  # Produit model removed

    result = []
    for code in all_codes:
        actual = sales_by_code.get(code, 0)
        objectif = obj_by_code.get(code, 0)
        nom = desc_by_code.get(code) or code
        pct = round(min(actual / objectif * 100, 100)) if objectif else 0
        result.append({
            "code_produit": code,
            "nom_produit": nom,
            "famille": '',  # Produit model removed
            "actual": round(actual),
            "objectif": round(objectif),
            "pct": pct,
        })

    return sorted(result, key=lambda x: (x['pct'], x['nom_produit']))
